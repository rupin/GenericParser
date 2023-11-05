import requests
from bs4 import BeautifulSoup
import openpyxl
from lxml import etree
from lxml import html
import re

def remove_html_tags(text):
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text)

def scrape_data(input_xlsx, url, output_xlsx):
    # Load the input Excel file
    workbook = openpyxl.load_workbook(input_xlsx)
    sheet = workbook.active

    # Create an output Excel file and add a new worksheet
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Write the header row to the output worksheet
    header_row = ["Question", "Option A", "Option B", "Option C", "Option D", "Answer", "Explanation"]
    output_sheet.append(header_row)

# Iterate through the rows in the input Excel file

#template
    rowObject = sheet.iter_cols(min_col=2, max_col=8, min_row=1, max_row= 1)  # +1 to adjust for 1-based indexing
    header_row = [cell.value for row in rowObject for cell in row]

    #template
    rowObject = sheet.iter_cols(min_col=2, max_col=8, min_row=2, max_row= 2)  # +1 to adjust for 1-based indexing
    xpathtemplate = [cell.value for row in rowObject for cell in row]


    #startIndex
    rowObject = sheet.iter_cols(min_col=2, max_col=8, min_row=3, max_row= 3)  # +1 to adjust for 1-based indexing
    startIndex = [int(cell.value) for row in rowObject for cell in row]

     #increments
    rowObject = sheet.iter_cols(min_col=2, max_col=8, min_row=4, max_row= 4)  # +1 to adjust for 1-based indexing
    increments = [int(cell.value) for row in rowObject for cell in row]

     #counts
    rowObject = sheet.iter_cols(min_col=2, max_col=8, min_row=5, max_row= 5)  # +1 to adjust for 1-based indexing
    counts = [int(cell.value) for row in rowObject for cell in row]

   # startIndex = sheet.row[2:]  # Extract XPath values for the first 7 columns
   # increment = sheet.row[3:]  # Extract XPath values for the first 7 columns
   # count = sheet.row[4:]  # Extract XPath values for the first 7 columns
   # question_data = []
    #print(header_row)
    
    #print(xpathtemplate)
    #print(startIndex)
    #print(increments)
    #print(counts)
    #exit()

    with open("mcq.html", encoding="utf8") as file:
        contents = file.read()
    soup = BeautifulSoup(contents,'html.parser')

    
        
    dom = etree.HTML(str(soup)) 
    #print(dom)
    question_data=[]
    for count in range(0, counts[0]):
        #print(count)
        question_data=[]
        for iteration in range(len(xpathtemplate)):
            # Find the HTML element using the XPath and extract its text
            #print(xpath)
            template=xpathtemplate[iteration]
            iterationStartIndex=startIndex[iteration]
            iterationincrement=increments[iteration]

            computedIndex=iterationStartIndex +iterationincrement*count
            computedPath=template.replace("#", str(computedIndex))

            element = dom.xpath(computedPath)
            #//*[@id="aspnetForm"]/div[4]/div[2]/p[8]/b
            #print(element)
            if element:
                text =  html.tostring(element[0], encoding="unicode")
                text=remove_html_tags(text)
            else:
                text = ""  # Handle the case where the element is not found
            question_data.append(text)
            #print(text)
        # Append the scraped data to the output worksheet
        output_sheet.append(question_data)  # Add the answer from the input file

    # Save the output Excel file
    output_workbook.save(output_xlsx)

if __name__ == "__main__":
    #input_xlsx = input("Enter the input Excel file name: ")
    #url = input("Enter the URL of the page to scrape: ")
    #output_xlsx = input("Enter the output Excel file name: ")
    input_xlsx="ParserData.xlsx"
    url="https://www.includehelp.com/data-analytics/mcq.aspx"
    output_xlsx="output.xlsx"

    scrape_data(input_xlsx, url, output_xlsx)
